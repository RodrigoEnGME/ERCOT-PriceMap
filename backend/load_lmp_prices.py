"""
Script para importar datos de precios LMP desde el archivo Excel.
Lee el archivo LPMs.xlsx y carga los precios en la base de datos.
"""
import sys
import os
from datetime import datetime
from sqlalchemy.orm import Session
import openpyxl

# Agregar el directorio raíz al path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.db.database import SessionLocal, engine, Base
from app.models import Node, PriceRecord

# Crear tablas si no existen
Base.metadata.create_all(bind=engine)


def load_prices_from_excel(excel_path: str, db: Session):
    """
    Carga los precios desde el archivo Excel a la base de datos.
    
    Formato esperado:
    - Columna A: Year
    - Columna B: Month  
    - Columna C: Hour
    - Columnas D-FF: Precios para nodos 1-150
    """
    print(f"Cargando archivo: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    
    print(f"Total de filas: {ws.max_row}")
    print(f"Total de columnas: {ws.max_column}")
    
    # Leer encabezados (primera fila)
    headers = [cell.value for cell in ws[1]]
    print(f"Encabezados: {headers[:10]}...")
    
    # Obtener mapeo de códigos de nodo a IDs de la base de datos
    nodes = db.query(Node).filter(Node.is_active == True).all()
    node_map = {}  # Mapea código de nodo (Node #1, Node #2, etc.) a ID de BD
    
    for node in nodes:
        # Extraer número del código (ej: "Node #1" -> 1)
        if node.code and '#' in node.code:
            try:
                node_num = int(node.code.split('#')[1].strip())
                node_map[node_num] = node.id
            except:
                pass
    
    print(f"Nodos encontrados en BD: {len(node_map)}")
    
    if not node_map:
        print("ERROR: No se encontraron nodos en la base de datos.")
        print("Ejecute primero load_real_nodes.py")
        return
    
    # Contadores
    records_created = 0
    records_skipped = 0
    errors = 0
    
    # Procesar cada fila de datos (saltando el encabezado)
    print("\nProcesando datos...")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row_idx % 50 == 0:
            print(f"Procesando fila {row_idx}/{ws.max_row}...")
        
        try:
            year = row[0]
            month = row[1]
            hour = row[2]
            
            # Validar datos básicos
            if not year or not month or hour is None:
                records_skipped += 1
                continue
            
            # Crear timestamp
            try:
                timestamp = datetime(
                    year=int(year),
                    month=int(month),
                    day=1,  # Día 1 por defecto
                    hour=int(hour)
                )
            except (ValueError, TypeError) as e:
                print(f"  Error en fila {row_idx}: timestamp inválido ({year}-{month} {hour}:00) - {e}")
                errors += 1
                continue
            
            # Procesar precios para cada nodo (columnas 4 en adelante)
            for node_num in range(1, 151):  # Nodos 1-150
                col_idx = 3 + node_num - 1  # Columna en Excel (0-indexed)
                
                if col_idx >= len(row):
                    break
                
                price_value = row[col_idx]
                
                # Saltar si no hay precio o es 0
                if price_value is None or price_value == 0:
                    continue
                
                # Verificar que el nodo existe en BD
                if node_num not in node_map:
                    continue
                
                node_id = node_map[node_num]
                
                # Crear registro de precio
                try:
                    price_record = PriceRecord(
                        node_id=node_id,
                        timestamp=timestamp,
                        price=float(price_value),
                        market='MDA',  # Day-Ahead Market
                        solar_capture=None,
                        wind_capture=None
                    )
                    db.add(price_record)
                    records_created += 1
                    
                    # Commit cada 1000 registros para evitar problemas de memoria
                    if records_created % 1000 == 0:
                        db.commit()
                        print(f"  Guardados {records_created} registros...")
                        
                except Exception as e:
                    print(f"  Error creando registro para nodo {node_num} en {timestamp}: {e}")
                    errors += 1
        
        except Exception as e:
            print(f"  Error procesando fila {row_idx}: {e}")
            errors += 1
            continue
    
    # Commit final
    try:
        db.commit()
        print("\n✓ Commit final exitoso")
    except Exception as e:
        print(f"\n✗ Error en commit final: {e}")
        db.rollback()
        return
    
    print(f"\n{'='*60}")
    print(f"Resumen de importación:")
    print(f"  Registros creados: {records_created}")
    print(f"  Registros saltados: {records_skipped}")
    print(f"  Errores: {errors}")
    print(f"{'='*60}")


def main():
    excel_path = r'c:\Desarrollo\EMI\ERCOT_Priceing_Dashboard\Fuentes\LPMs.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"ERROR: Archivo no encontrado: {excel_path}")
        return
    
    print("Iniciando importación de precios LMP...")
    print("ADVERTENCIA: Este proceso puede tomar varios minutos.\n")
    
    db = SessionLocal()
    try:
        load_prices_from_excel(excel_path, db)
    except Exception as e:
        print(f"\nError fatal: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()
    
    print("\nProceso completado.")


if __name__ == "__main__":
    main()
