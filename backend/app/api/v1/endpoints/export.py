from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.db.database import get_db
from app.models import PriceRecord, Node, User, UserRole
from app.schemas import ExportRequest, DataType
from app.api.dependencies import get_current_active_user

router = APIRouter(prefix="/export", tags=["Export"])


@router.post("/excel")
async def export_to_excel(
    export_data: ExportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Export price data to Excel file."""
    # Check user permissions (only premium and admin can export)
    if current_user.role == UserRole.BASIC:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Export feature requires premium or admin account"
        )
    
    # Get nodes
    nodes = db.query(Node).filter(Node.id.in_(export_data.node_ids)).all()
    if not nodes:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No nodes found"
        )
    
    # Select data field
    field_map = {
        DataType.PRICE: PriceRecord.price,
        DataType.SOLAR_CAPTURE: PriceRecord.solar_capture,
        DataType.WIND_CAPTURE: PriceRecord.wind_capture
    }
    data_field = field_map[export_data.data_type]
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Create data sheet
    ws_data = wb.active
    ws_data.title = "Data"
    
    # Headers
    headers = ["Timestamp", "Node Code", "Node Name", export_data.data_type.value.replace('_', ' ').title()]
    ws_data.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws_data[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Fetch and add data
    for node in nodes:
        records = (
            db.query(PriceRecord.timestamp, data_field)
            .filter(
                and_(
                    PriceRecord.node_id == node.id,
                    PriceRecord.timestamp >= export_data.start_date,
                    PriceRecord.timestamp <= export_data.end_date,
                    data_field.isnot(None)
                )
            )
            .order_by(PriceRecord.timestamp)
            .all()
        )
        
        for record in records:
            ws_data.append([
                record[0].strftime("%Y-%m-%d %H:%M:%S"),
                node.code,
                node.name,
                float(record[1]) if record[1] else None
            ])
    
    # Auto-adjust column widths
    for column in ws_data.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_data.column_dimensions[column_letter].width = adjusted_width
    
    # Create aggregations sheet if requested
    if export_data.include_aggregations:
        ws_agg = wb.create_sheet("Aggregations")
        ws_agg.append(["Node Code", "Node Name", "Average", "Maximum", "Minimum", "Count"])
        
        # Style headers
        for cell in ws_agg[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for node in nodes:
            from sqlalchemy import func
            stats = (
                db.query(
                    func.avg(data_field).label('avg'),
                    func.max(data_field).label('max'),
                    func.min(data_field).label('min'),
                    func.count(data_field).label('count')
                )
                .filter(
                    and_(
                        PriceRecord.node_id == node.id,
                        PriceRecord.timestamp >= export_data.start_date,
                        PriceRecord.timestamp <= export_data.end_date,
                        data_field.isnot(None)
                    )
                )
                .first()
            )
            
            ws_agg.append([
                node.code,
                node.name,
                float(stats[0]) if stats[0] else None,
                float(stats[1]) if stats[1] else None,
                float(stats[2]) if stats[2] else None,
                stats[3]
            ])
        
        # Auto-adjust column widths
        for column in ws_agg.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_agg.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename
    filename = f"ercot_data_{export_data.data_type.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
